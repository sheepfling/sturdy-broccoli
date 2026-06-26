from __future__ import annotations

import csv
import json
from pathlib import Path

from openpyxl import load_workbook

from scripts.generate_requirement_compliance_spreadsheets import generate_exports


ROOT = Path(__file__).resolve().parents[2]


def test_requirement_compliance_spreadsheet_export_writes_both_editions(tmp_path: Path) -> None:
    created = generate_exports(tmp_path)
    names = {path.name for path in created}

    assert names == {
        "requirements_2010_backend_compliance.xlsx",
        "requirements_2010_backend_compliance_detail.csv",
        "requirements_2010_backend_compliance_policy_parents.csv",
        "requirements_2010_backend_compliance_summary.csv",
        "requirements_2025_backend_compliance.xlsx",
        "requirements_2025_backend_compliance_detail.csv",
        "requirements_2025_backend_compliance_summary.csv",
    }

    workbook_2010 = load_workbook(tmp_path / "requirements_2010_backend_compliance.xlsx", read_only=True)
    workbook_2025 = load_workbook(tmp_path / "requirements_2025_backend_compliance.xlsx", read_only=True)
    assert workbook_2010.sheetnames == ["summary", "detail", "metadata", "policy_parents"]
    assert workbook_2025.sheetnames == ["summary", "detail", "metadata"]

    with (tmp_path / "requirements_2010_backend_compliance_detail.csv").open(newline="", encoding="utf-8") as handle:
        rows_2010 = list(csv.DictReader(handle))
    with (tmp_path / "requirements_2010_backend_compliance_policy_parents.csv").open(newline="", encoding="utf-8") as handle:
        policy_rows_2010 = list(csv.DictReader(handle))
    with (tmp_path / "requirements_2025_backend_compliance_detail.csv").open(newline="", encoding="utf-8") as handle:
        rows_2025 = list(csv.DictReader(handle))
    metadata_2025 = list(
        load_workbook(tmp_path / "requirements_2025_backend_compliance.xlsx", read_only=True)["metadata"].iter_rows(values_only=True)
    )
    metadata_map_2025 = {
        str(field): str(value)
        for field, value in metadata_2025[1:]
        if field is not None
    }
    metadata_2010 = list(
        load_workbook(tmp_path / "requirements_2010_backend_compliance.xlsx", read_only=True)["metadata"].iter_rows(values_only=True)
    )
    metadata_map_2010 = {
        str(field): str(value)
        for field, value in metadata_2010[1:]
        if field is not None
    }
    summary_2010 = {}
    with (tmp_path / "requirements_2010_backend_compliance_summary.csv").open(newline="", encoding="utf-8") as handle:
        for row in csv.DictReader(handle):
            summary_2010[row["metric"]] = row["value"]
    summary_2025 = {}
    with (tmp_path / "requirements_2025_backend_compliance_summary.csv").open(newline="", encoding="utf-8") as handle:
        for row in csv.DictReader(handle):
            summary_2025[row["metric"]] = row["value"]

    assert any(row["requirement_id"] == "HLA1516.1-FM-4.1.5-001" for row in rows_2010)
    assert any(row["service_group"] == "Data distribution management" for row in rows_2025)
    assert any(row["canonical_status"] == "partial" for row in rows_2010)
    assert any(row["canonical_disposition"] == "covered" for row in rows_2025)
    assert any(
        row["requirement_id"] == "HLA1516.1-FM-4.1.5-001"
        and row["python_runtime_disposition"] == "verified"
        and row["pitch_runtime_disposition"] == "blocked"
        for row in rows_2010
    )
    assert any(
        row["requirement_id"] == "REQ-OMT-SCHEMA-001"
        and row["python_runtime_disposition"] == "not-applicable"
        and row["pitch_runtime_disposition"] == "classification-required"
        for row in rows_2010
    )
    assert any(
        row["requirement_id"] == "HLA1516.1-FM-4.10-TEST-001"
        and "hosted gRPC and REST replay recorded separately" in row["title"]
        and "The direct Python 2010 lane is proven directly" in row["notes"]
        and "hosted gRPC and REST lifecycle replay is recorded separately" in row["notes"]
        for row in rows_2010
    )
    assert any(
        row["requirement_id"] == "HLA1516.1-FM-4.11-TEST-001"
        and "hosted gRPC and REST replay recorded separately" in row["title"]
        and "hosted gRPC and REST synchronization replay is recorded separately" in row["notes"]
        for row in rows_2010
    )
    assert any(
        row["requirement_id"] == "HLA1516.1-FM-4.16-TEST-001"
        and "hosted gRPC and REST replay recorded separately" in row["title"]
        and "hosted gRPC and REST save/restore replay is recorded separately" in row["notes"]
        for row in rows_2010
    )
    execution_rows_2010 = {
        row["requirement_id"]: row
        for row in rows_2010
        if row["requirement_id"] in {
            "HLA1516.1-FM-4.1-001",
            "HLA1516.1-FM-4.6-001",
            "HLA1516.1-FM-4.9-001",
            "HLA1516.1-FM-4.10-001",
            "HLA1516.1-OM-6.8-001",
            "HLA1516.1-OM-6.10-001",
            "HLA1516.1-OM-6.12-001",
            "HLA1516.1-OM-6.14-001",
            "HLA1516.1-OM-6.16-001",
            "HLA1516.1-OM-6.19-001",
            "HLA1516.1-OM-6.25-001",
            "HLA1516.1-DDM-9.12-001",
            "HLA1516.1-DDM-9.13-001",
        }
    }
    assert set(execution_rows_2010) == {
        "HLA1516.1-FM-4.1-001",
        "HLA1516.1-FM-4.6-001",
        "HLA1516.1-FM-4.9-001",
        "HLA1516.1-FM-4.10-001",
        "HLA1516.1-OM-6.8-001",
        "HLA1516.1-OM-6.10-001",
        "HLA1516.1-OM-6.12-001",
        "HLA1516.1-OM-6.14-001",
        "HLA1516.1-OM-6.16-001",
        "HLA1516.1-OM-6.19-001",
        "HLA1516.1-OM-6.25-001",
        "HLA1516.1-DDM-9.12-001",
        "HLA1516.1-DDM-9.13-001",
    }
    assert execution_rows_2010["HLA1516.1-FM-4.6-001"]["canonical_status"] == "pass"
    assert execution_rows_2010["HLA1516.1-FM-4.6-001"]["python_runtime_disposition"] == "verified"
    assert execution_rows_2010["HLA1516.1-FM-4.9-001"]["title"] == "RTI shall provide Join Federation Execution"
    assert execution_rows_2010["HLA1516.1-FM-4.10-001"]["title"] == "RTI shall provide Resign Federation Execution"
    assert execution_rows_2010["HLA1516.1-OM-6.10-001"]["title"] == "RTI shall provide Update Attribute Values."
    assert execution_rows_2010["HLA1516.1-OM-6.12-001"]["title"] == "RTI shall provide Send Interaction."
    assert execution_rows_2010["HLA1516.1-OM-6.25-001"]["canonical_status"] == "partial"
    assert execution_rows_2010["HLA1516.1-OM-6.25-001"]["python_runtime_disposition"] == "verified"
    assert "supported reliable plus best-effort subset" in execution_rows_2010["HLA1516.1-OM-6.25-001"]["notes"]
    assert execution_rows_2010["HLA1516.1-DDM-9.12-001"]["title"] == "RTI shall route interactions based on region overlap where dimensions apply"
    assert execution_rows_2010["HLA1516.1-DDM-9.13-001"]["title"] == "RTI shall route attribute-value update requests based on region overlap"
    assert not any("transport-equivalence checks across native, gRPC, and REST entry points" in row["title"] for row in rows_2010)
    assert not any("Native lifecycle tests now have explicit hosted gRPC and REST lifecycle smoke coverage." in row["notes"] for row in rows_2010)
    assert len(policy_rows_2010) == 12
    assert sum(int(row["supported_subset_pass_count"]) for row in policy_rows_2010) == 26
    assert all(row["supported_subset_links_match"] == "yes" for row in policy_rows_2010)
    assert any(
        row["broad_requirement_id"] == "HLA1516.1-DM-5.1.6-001"
        and row["supported_subset_pass_ids"] == "HLA1516.1-DM-5.1.6-002"
        and row["policy_basis"] == "logical-time-update-rate-only"
        for row in policy_rows_2010
    )
    assert any(
        row["broad_requirement_id"] == "HLA1516.1-OM-6.24-001"
        and row["supported_subset_pass_ids"] == "HLA1516.1-OM-6.24-002; HLA1516.1-OM-6.24-003; HLA1516.1-OM-6.24-004"
        and row["policy_basis"] == "reliable-best-effort-transport-only"
        for row in policy_rows_2010
    )
    assert any(row["canonical_disposition"] == "duplicate/umbrella" for row in rows_2025)
    execution_groups_2025 = [
        row for row in rows_2025
        if row["closure_wave"] == "1-fi-service-and-binding-disposition"
        and row["service_group"] in {
            "Federation management",
            "Object management",
            "Data distribution management",
        }
    ]
    assert execution_groups_2025
    assert any(
        row["service_group"] == "Federation management"
        and row["canonical_disposition"] == "covered"
        and "hla_2025_requirement_disposition_ledger.csv" in row["backend_resolution_reference"]
        and "hla_2025_fi_binding_surface_matrix.csv" in row["backend_resolution_reference"]
        for row in execution_groups_2025
    )
    assert any(
        row["service_group"] == "Object management"
        and row["canonical_disposition"] == "covered"
        and row["python_runtime_resolution"] == "grouped row set already has direct python1516_2025 evidence; keep per-service truth in the harmonization ledger"
        for row in execution_groups_2025
    )
    assert any(
        row["service_group"] == "Data distribution management"
        and row["canonical_disposition"] == "covered"
        and row["hosted_fedpro_resolution"] == "FedPro route/support is tracked separately in the FI binding matrix and harmonization ledger"
        for row in execution_groups_2025
    )
    assert any(
        row["canonical_disposition"] == "duplicate/umbrella"
        and "not a standalone runtime claim" in row["python_runtime_resolution"]
        and row["acceptance_gate"]
        == "Every row in this group has explicit owner-doc evidence, row-level disposition, child-row or backend-resolution references, and promotion/no-promote semantics reviewed."
        for row in rows_2025
    )
    assert any(
        row["canonical_disposition"] == "retired/legacy-only"
        and "not active runtime ownership" in row["python_runtime_resolution"]
        and "not an active hosted-route claim" in row["hosted_fedpro_resolution"]
        and row["acceptance_gate"]
        == "Every row in this group has explicit exclusion-owner evidence, replacement mapping, row-level disposition, and compatibility-only promotion semantics reviewed."
        for row in rows_2025
    )
    assert json.loads(summary_2010["python_runtime_disposition_counts"]).get("verified", 0) > 0
    assert json.loads(summary_2010["pitch_runtime_disposition_counts"]).get("classification-required", 0) > 0
    assert summary_2010["policy_parent_partial_count"] == "12"
    assert summary_2010["policy_parent_supported_subset_pass_count"] == "26"
    assert metadata_map_2010["canonical_status_meaning"].startswith("Requirement coverage state only")
    assert "direct Python 2010 runtime lane" in metadata_map_2010["python_runtime_disposition_meaning"]
    assert "Hosted gRPC/REST replay" in metadata_map_2010["python_runtime_disposition_meaning"]
    assert "Pitch is independently verified" in metadata_map_2010["pitch_runtime_disposition_meaning"]
    assert "not an open Python gap list" in metadata_map_2010["policy_parents_meaning"]
    assert summary_2025["tracked_row_universe"] == "691"
    assert summary_2025["active_normative_non_retired_non_umbrella_denominator"] == "645"
    assert summary_2025["active_normative_direct_coverage_count"] == "645"
    assert summary_2025["active_normative_direct_coverage_fraction"] == "645 / 645 = 100%"
    assert summary_2025["duplicate_umbrella_row_count"] == "22"
    assert summary_2025["retired_legacy_only_row_count"] == "24"
    assert summary_2025["tracked_rows_outside_active_direct_support_denominator"] == "46"
    assert summary_2025["row_level_source_artifact"] == "requirements/2025/harmonization/hla_2025_requirement_disposition_ledger.csv"
    assert metadata_map_2025["canonical_disposition_meaning"].startswith("Requirement coverage state only")
    assert metadata_map_2025["row_level_owner_source"] == "requirements/2025/harmonization/hla_2025_requirement_disposition_ledger.csv"
    assert "100% dispositioned across all 691 tracked rows" in metadata_map_2025["denominator_rule"]
    assert "645 active normative non-retired non-umbrella rows" in metadata_map_2025["denominator_rule"]
    assert "do not restate this grouped packet as 691 / 691 covered" in metadata_map_2025["denominator_rule"]
    assert "64 grouped buckets" in metadata_map_2025["grouped_packet_scope"]
    assert "691-row canonical denominator" in metadata_map_2025["grouped_packet_scope"]
    assert "bounded hosted-route surface over python1516_2025" in metadata_map_2025["hosted_fedpro_resolution_meaning"]
    assert "Pitch proto HLA 4 / 202X overlap" in metadata_map_2025["pitch_202x_resolution_meaning"]


def test_requirement_compliance_export_doc_keeps_2025_denominator_split_explicit() -> None:
    text = (
        ROOT
        / "docs"
        / "verification"
        / "requirement_compliance_exports.md"
    ).read_text(encoding="utf-8")
    normalized = " ".join(text.split())

    assert "`100% dispositioned` across all `691` tracked rows" in text
    assert "`100% covered` across the `645` active normative non-retired" in text
    assert "`duplicate/umbrella`" in text
    assert "`retired/legacy-only`" in text
    assert "generated presentation surface, not a checked-in owner surface" in normalized
    assert "`analysis/compliance/presentation_packets/` is ignored by git" in normalized
    assert "`canonical_disposition` answers coverage state only" in text
    assert "`canonical_status` answers requirement coverage state only" in text
    assert "hosted `gRPC`/`REST` replay" in text
    assert "`hosted_fedpro_resolution` answers whether FedPro is a bounded hosted-route" in text
    assert "Pitch proto HLA 4 / `202X`" in text
    assert "`analysis/compliance/compliance.before/`" in text
    assert "as the current owner surface or the current spreadsheet handoff packet" in normalized
    assert "2010_python_rti_bounded_family_execution_worklist.md" in text
    assert "`requirements_2010_backend_compliance_policy_parents.csv`" in text
    assert "`policy_parents`" in text
    assert "12 broad partial rows" in text
    assert "not an open Python gap list" in text


def test_pitch_202x_group_resolution_and_presentation_packet_stay_aligned() -> None:
    pitch_group_rows = list(
        csv.DictReader(
            (
                ROOT
                / "requirements"
                / "2025"
                / "harmonization"
                / "hla_2025_pitch_202x_group_resolution.csv"
            ).open(newline="", encoding="utf-8")
        )
    )
    presentation_rows = list(
        csv.DictReader(
            (
                ROOT
                / "analysis"
                / "compliance"
                / "presentation_packets"
                / "requirements_2025_backend_compliance_detail.csv"
            ).open(newline="", encoding="utf-8")
        )
    )

    assert len(pitch_group_rows) == 64

    def _pitch_key(row: dict[str, str]) -> tuple[str, ...]:
        return (
            row["closure_wave"],
            row["priority"],
            row["area"],
            row["service_group"],
            row["canonical_disposition"],
            row["pitch_202x_resolution"],
            row["pitch_202x_owner_doc"],
            row["pitch_202x_evidence_packet"],
            row["pitch_202x_primary_command"],
            row["pitch_202x_scope_note"],
        )

    presentation_counter = {}
    for row in presentation_rows:
        pitch_row_key = (
            row["closure_wave"],
            row["priority"],
            row["area"],
            row["service_group"],
            row["canonical_disposition"],
            row["pitch_202x_resolution"],
            "docs/requirements/ieee-1516-2025/pitch_202x_bounded_comparison.md",
            "",
            "",
            "",
        )
        # We separately assert the owner doc and evidence packet references below.
        presentation_counter[pitch_row_key[:6]] = presentation_counter.get(pitch_row_key[:6], 0) + 1

    expected_counter = {}
    for row in pitch_group_rows:
        reduced_key = _pitch_key(row)[:6]
        expected_counter[reduced_key] = expected_counter.get(reduced_key, 0) + 1

    assert presentation_counter == expected_counter

    for row in presentation_rows:
        matching_pitch_rows = [
            pitch_row
            for pitch_row in pitch_group_rows
            if (
                pitch_row["closure_wave"],
                pitch_row["priority"],
                pitch_row["area"],
                pitch_row["service_group"],
                pitch_row["canonical_disposition"],
                pitch_row["pitch_202x_resolution"],
            )
            == (
                row["closure_wave"],
                row["priority"],
                row["area"],
                row["service_group"],
                row["canonical_disposition"],
                row["pitch_202x_resolution"],
            )
        ]
        assert matching_pitch_rows
        if matching_pitch_rows[0]["pitch_202x_evidence_packet"].strip():
            for expected_ref in (
                "docs/requirements/ieee-1516-2025/pitch_202x_bounded_comparison.md",
                "requirements/2025/harmonization/hla_2025_pitch_202x_group_resolution.csv",
            ):
                assert expected_ref in row["backend_resolution_reference"]
        if matching_pitch_rows[0]["pitch_202x_evidence_packet"].strip():
            for expected_ref in matching_pitch_rows[0]["pitch_202x_evidence_packet"].split("; "):
                assert expected_ref in row["backend_resolution_reference"]


def test_pitch_202x_group_resolution_keeps_bounded_vendor_reading_explicit_by_disposition() -> None:
    rows = list(
        csv.DictReader(
            (
                ROOT
                / "requirements"
                / "2025"
                / "harmonization"
                / "hla_2025_pitch_202x_group_resolution.csv"
            ).open(newline="", encoding="utf-8")
        )
    )

    assert all(
        row["pitch_202x_owner_doc"] == "docs/requirements/ieee-1516-2025/pitch_202x_bounded_comparison.md"
        for row in rows
    )
    assert all(row["pitch_202x_scope_note"].strip() for row in rows)

    covered_rows = [row for row in rows if row["canonical_disposition"] == "covered"]
    umbrella_rows = [row for row in rows if row["canonical_disposition"] == "duplicate/umbrella"]
    retired_rows = [row for row in rows if row["canonical_disposition"] == "retired/legacy-only"]

    assert len(covered_rows) == 57
    assert len(umbrella_rows) == 5
    assert len(retired_rows) == 2

    assert any(
        row["pitch_202x_resolution"]
        == "vendor-branded proto HLA 4 / 202X surface may overlap this FI bucket, but grouped Pitch parity still has to be read from linked owner artifacts rather than inferred here"
        and row["pitch_202x_primary_command"] == "./tools/pitch 202x-micro-certify"
        and "does not prove clause-complete IEEE 1516.1-2025 vendor conformance" in row["pitch_202x_scope_note"]
        for row in covered_rows
    )
    assert any(
        row["pitch_202x_resolution"]
        == "n/a at grouped service-usage level unless a mirrored FI owner artifact explicitly records Pitch proto HLA 4 / 202X support or divergence"
        and row["pitch_202x_primary_command"] == ""
        and row["pitch_202x_evidence_packet"] == ""
        for row in covered_rows
    )
    assert any(
        row["pitch_202x_resolution"]
        == "n/a: this grouped OMT bucket is parser or validator owned, not a Pitch proto HLA 4 / 202X runtime claim"
        and row["pitch_202x_primary_command"] == ""
        and row["pitch_202x_evidence_packet"] == ""
        for row in covered_rows
    )
    assert all(
        row["pitch_202x_resolution"]
        == "umbrella only; any Pitch proto HLA 4 / 202X claim must resolve through child FI rows and binding boundary notes"
        and row["pitch_202x_primary_command"] == "./tools/pitch 202x-micro-certify"
        and "Use this only as a grouped callback/binding interpretation surface." in row["pitch_202x_scope_note"]
        for row in umbrella_rows
        if row["area"] == "Callback/configuration/binding deltas"
    )
    assert all(
        row["pitch_202x_resolution"]
        == "n/a at framework umbrella level; use child FI or OMT owner artifacts if Pitch proto HLA 4 / 202X becomes relevant"
        and row["pitch_202x_primary_command"] == ""
        and row["pitch_202x_evidence_packet"] == ""
        for row in umbrella_rows
        if row["area"] == "Framework and Rules"
    )
    assert all(
        row["pitch_202x_resolution"]
        == "legacy mapping only; not an active Pitch proto HLA 4 / 202X behavior-support claim"
        and row["pitch_202x_primary_command"] == ""
        and row["pitch_202x_evidence_packet"] == ""
        and row["pitch_202x_scope_note"] == "See owner doc."
        for row in retired_rows
    )


def test_2025_harmonization_worklist_keeps_non_covered_acceptance_gates_and_framework_closeout_honest() -> None:
    rows = list(
        csv.DictReader(
            (
                ROOT
                / "requirements"
                / "2025"
                / "harmonization"
                / "hla_2025_harmonization_worklist.csv"
            ).open(newline="", encoding="utf-8")
        )
    )

    umbrella_rows = [row for row in rows if row["canonical_disposition"] == "duplicate/umbrella"]
    retired_rows = [row for row in rows if row["canonical_disposition"] == "retired/legacy-only"]

    assert len(umbrella_rows) == 5
    assert len(retired_rows) == 2

    assert all(
        row["acceptance_gate"]
        == "Every row in this group has explicit exclusion-owner evidence, replacement mapping, row-level disposition, and compatibility-only promotion semantics reviewed."
        for row in retired_rows
    )
    assert all(
        row["acceptance_gate"]
        == "Every row in this group has explicit owner-doc evidence, row-level disposition, child-row or backend-resolution references, and promotion/no-promote semantics reviewed."
        for row in umbrella_rows
    )
    assert all(
        row["closure_goal"]
        == "Keep the child-row map and framework owner reading synchronized; do not relabel parent normalization rows as standalone proof without a narrower direct claim."
        for row in umbrella_rows
        if row["area"] == "Framework and Rules"
    )
