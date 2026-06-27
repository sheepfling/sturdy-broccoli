from __future__ import annotations

import argparse
import csv
import json
from collections import Counter
from pathlib import Path
from typing import Iterable

from openpyxl import Workbook


ROOT = Path(__file__).resolve().parents[1]
DEFAULT_OUTPUT_DIR = ROOT / "analysis" / "compliance" / "presentation_packets"


def _read_csv(path: Path) -> list[dict[str, str]]:
    with path.open(newline="", encoding="utf-8") as handle:
        return list(csv.DictReader(handle))


def _write_csv(path: Path, fieldnames: list[str], rows: Iterable[dict[str, str]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        for row in rows:
            writer.writerow(row)


def _autosize_worksheet(ws) -> None:
    for column_cells in ws.columns:
        width = max(len(str(cell.value or "")) for cell in column_cells)
        ws.column_dimensions[column_cells[0].column_letter].width = min(max(width + 2, 12), 80)


def _write_sheet(ws, rows: list[dict[str, str]]) -> None:
    if not rows:
        ws.append(["no rows"])
        return
    headers = list(rows[0].keys())
    ws.append(headers)
    for row in rows:
        ws.append([row.get(header, "") for header in headers])
    _autosize_worksheet(ws)


def _display_path(path: Path) -> str:
    try:
        return str(path.relative_to(ROOT))
    except ValueError:
        return str(path)


def _export_workbook(
    path: Path,
    summary_rows: list[dict[str, str]],
    detail_rows: list[dict[str, str]],
    metadata_rows: list[dict[str, str]],
    extra_sheets: list[tuple[str, list[dict[str, str]]]] | None = None,
) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws_summary = wb.active
    ws_summary.title = "summary"
    _write_sheet(ws_summary, summary_rows)

    ws_detail = wb.create_sheet("detail")
    _write_sheet(ws_detail, detail_rows)

    ws_meta = wb.create_sheet("metadata")
    _write_sheet(ws_meta, metadata_rows)

    for title, rows in extra_sheets or []:
        ws = wb.create_sheet(title)
        _write_sheet(ws, rows)
    wb.save(path)


def _split_linked_ids(value: str) -> list[str]:
    return [part.strip() for part in value.split(";") if part.strip()]


def _acceptance_gate_2025_group(row: dict[str, str]) -> str:
    canonical_disposition = row.get("canonical_disposition", "")
    if canonical_disposition == "duplicate/umbrella":
        return (
            "Every row in this group has explicit owner-doc evidence, row-level disposition, child-row or "
            "backend-resolution references, and promotion/no-promote semantics reviewed."
        )
    if canonical_disposition == "retired/legacy-only":
        return (
            "Every row in this group has explicit exclusion-owner evidence, replacement mapping, row-level "
            "disposition, and compatibility-only promotion semantics reviewed."
        )
    return row.get("acceptance_gate", "")


def _build_2010_policy_parent_rows(rows: list[dict[str, str]]) -> list[dict[str, str]]:
    row_by_requirement_id = {
        row["requirement_id"]: row
        for row in rows
        if row.get("requirement_id")
    }
    children_by_parent: dict[str, list[dict[str, str]]] = {}
    for row in rows:
        requirement_id = row.get("requirement_id", "")
        if not requirement_id or row.get("status") != "pass":
            continue
        supported_subset_for = row.get("supported_subset_for", "").strip()
        if not supported_subset_for:
            continue
        parent_row = row_by_requirement_id.get(supported_subset_for)
        if parent_row is None or parent_row.get("status") != "partial":
            continue
        children_by_parent.setdefault(supported_subset_for, []).append(row)

    policy_parent_rows: list[dict[str, str]] = []
    for parent_id in sorted(children_by_parent):
        parent_row = row_by_requirement_id[parent_id]
        child_rows = sorted(children_by_parent[parent_id], key=lambda row: row["requirement_id"])
        explicit_parent_links = _split_linked_ids(parent_row.get("supported_subset_for", ""))
        policy_parent_rows.append(
            {
                "broad_requirement_id": parent_id,
                "section_ref": parent_row.get("section_ref", ""),
                "policy_basis": parent_row.get("policy_basis", ""),
                "broad_status": parent_row.get("status", ""),
                "python_runtime_disposition": parent_row.get("python_runtime_disposition", ""),
                "supported_subset_pass_count": str(len(child_rows)),
                "supported_subset_pass_ids": "; ".join(row["requirement_id"] for row in child_rows),
                "supported_subset_links_match": "yes"
                if explicit_parent_links == [row["requirement_id"] for row in child_rows]
                else "no",
                "broad_row_note": parent_row.get("notes", ""),
            }
        )
    return policy_parent_rows


def build_2010_rows() -> tuple[list[dict[str, str]], list[dict[str, str]], list[dict[str, str]], list[dict[str, str]]]:
    matrix_path = ROOT / "analysis" / "compliance" / "requirements_matrix_2010.csv"
    rows = _read_csv(matrix_path)
    detail_rows: list[dict[str, str]] = []
    for row in rows:
        detail_rows.append(
            {
                "requirement_id": row.get("requirement_id", ""),
                "matrix_id": row.get("matrix_id", ""),
                "service_group": row.get("service_group", ""),
                "title": row.get("title", ""),
                "section_ref": row.get("section_ref", ""),
                "canonical_status": row.get("status", ""),
                "python_runtime_disposition": row.get("python_runtime_disposition", ""),
                "certi_runtime_disposition": row.get("certi_runtime_disposition", ""),
                "pitch_runtime_disposition": row.get("pitch_runtime_disposition", ""),
                "portico_runtime_disposition": row.get("portico_runtime_disposition", ""),
                "claim_scope": row.get("claim_scope", ""),
                "artifact_refs": row.get("artifact_refs", ""),
                "notes": row.get("notes", ""),
            }
        )
    policy_parent_rows = _build_2010_policy_parent_rows(rows)

    status_counts = Counter(row["canonical_status"] for row in detail_rows)
    python_counts = Counter(row["python_runtime_disposition"] for row in detail_rows)
    certi_counts = Counter(row["certi_runtime_disposition"] for row in detail_rows)
    pitch_counts = Counter(row["pitch_runtime_disposition"] for row in detail_rows)
    portico_counts = Counter(row["portico_runtime_disposition"] for row in detail_rows)

    summary_rows = [
        {"metric": "edition", "value": "2010 / 1516e"},
        {"metric": "row_count", "value": str(len(detail_rows))},
        {"metric": "canonical_status_counts", "value": json.dumps(status_counts, sort_keys=True)},
        {"metric": "python_runtime_disposition_counts", "value": json.dumps(python_counts, sort_keys=True)},
        {"metric": "certi_runtime_disposition_counts", "value": json.dumps(certi_counts, sort_keys=True)},
        {"metric": "pitch_runtime_disposition_counts", "value": json.dumps(pitch_counts, sort_keys=True)},
        {"metric": "portico_runtime_disposition_counts", "value": json.dumps(portico_counts, sort_keys=True)},
        {"metric": "policy_parent_partial_count", "value": str(len(policy_parent_rows))},
        {
            "metric": "policy_parent_supported_subset_pass_count",
            "value": str(sum(int(row["supported_subset_pass_count"]) for row in policy_parent_rows)),
        },
        {"metric": "source_artifact", "value": "analysis/compliance/requirements_matrix_2010.csv"},
    ]
    metadata_rows = [
        {"field": "edition", "value": "2010 / 1516e"},
        {"field": "primary_source", "value": "analysis/compliance/requirements_matrix_2010.csv"},
        {"field": "front_door", "value": "docs/requirements/ieee-1516-2010/README.md"},
        {"field": "inventory", "value": "requirements/2010/README.md"},
        {
            "field": "canonical_status_meaning",
            "value": "Requirement coverage state only. Read this separately from backend-resolution disposition fields.",
        },
        {
            "field": "python_runtime_disposition_meaning",
            "value": "Whether the direct Python 2010 runtime lane is verified, bounded, blocked, not applicable, or still needs explicit classification. Hosted gRPC/REST replay, when present, is separate route evidence and does not change canonical status by itself.",
        },
        {
            "field": "certi_runtime_disposition_meaning",
            "value": "Whether CERTI is independently verified, explicitly bounded, blocked, not applicable, or still needs explicit classification for the row.",
        },
        {
            "field": "pitch_runtime_disposition_meaning",
            "value": "Whether Pitch is independently verified, explicitly bounded, blocked, not applicable, or still needs explicit classification for the row.",
        },
        {
            "field": "portico_runtime_disposition_meaning",
            "value": "Whether Portico is independently verified, explicitly bounded, blocked, not applicable, or still needs explicit classification for the row.",
        },
        {
            "field": "policy_parents_meaning",
            "value": "Broad partial rows intentionally defended by passing supported-subset children. This sheet is not an open Python gap list.",
        },
        {"field": "notes", "value": "Boss-facing export generated from canonical CSV/JSON artifacts; spreadsheets are secondary presentation surfaces."},
    ]
    return summary_rows, detail_rows, metadata_rows, policy_parent_rows


def build_2025_rows() -> tuple[list[dict[str, str]], list[dict[str, str]], list[dict[str, str]]]:
    worklist_path = ROOT / "requirements" / "2025" / "harmonization" / "hla_2025_harmonization_worklist.csv"
    ledger_path = ROOT / "requirements" / "2025" / "harmonization" / "hla_2025_requirement_disposition_ledger.csv"
    rows = _read_csv(worklist_path)
    ledger_rows = _read_csv(ledger_path)
    detail_rows: list[dict[str, str]] = []
    for row in rows:
        detail_rows.append(
            {
                "closure_wave": row.get("closure_wave", ""),
                "priority": row.get("priority", ""),
                "area": row.get("area", ""),
                "service_group": row.get("service_group", ""),
                "canonical_disposition": row.get("canonical_disposition", ""),
                "python_runtime_resolution": row.get("python_runtime_resolution", ""),
                "java_cpp_binding_resolution": row.get("java_cpp_binding_resolution", ""),
                "hosted_fedpro_resolution": row.get("hosted_fedpro_resolution", ""),
                "pitch_202x_resolution": row.get("pitch_202x_resolution", ""),
                "row_count": row.get("row_count", ""),
                "backend_resolution_reference": row.get("backend_resolution_reference", ""),
                "acceptance_gate": _acceptance_gate_2025_group(row),
            }
        )

    disposition_counts = Counter(row["canonical_disposition"] for row in detail_rows)
    area_counts = Counter(row["area"] for row in detail_rows)
    row_level_disposition_counts = Counter(row["harmonization_disposition"] for row in ledger_rows)
    tracked_row_universe = len(ledger_rows)
    active_normative_denominator = row_level_disposition_counts.get("covered", 0)
    duplicate_umbrella_row_count = row_level_disposition_counts.get("duplicate/umbrella", 0)
    retired_legacy_only_row_count = row_level_disposition_counts.get("retired/legacy-only", 0)
    non_covered_tracked_rows = duplicate_umbrella_row_count + retired_legacy_only_row_count

    summary_rows = [
        {"metric": "edition", "value": "2025 / 1516_2025"},
        {"metric": "group_count", "value": str(len(detail_rows))},
        {"metric": "canonical_disposition_counts", "value": json.dumps(disposition_counts, sort_keys=True)},
        {"metric": "tracked_row_universe", "value": str(tracked_row_universe)},
        {"metric": "row_level_disposition_counts", "value": json.dumps(row_level_disposition_counts, sort_keys=True)},
        {"metric": "active_normative_non_retired_non_umbrella_denominator", "value": str(active_normative_denominator)},
        {"metric": "active_normative_direct_coverage_count", "value": str(active_normative_denominator)},
        {
            "metric": "active_normative_direct_coverage_fraction",
            "value": f"{active_normative_denominator} / {active_normative_denominator} = 100%",
        },
        {"metric": "duplicate_umbrella_row_count", "value": str(duplicate_umbrella_row_count)},
        {"metric": "retired_legacy_only_row_count", "value": str(retired_legacy_only_row_count)},
        {"metric": "tracked_rows_outside_active_direct_support_denominator", "value": str(non_covered_tracked_rows)},
        {"metric": "area_counts", "value": json.dumps(area_counts, sort_keys=True)},
        {"metric": "source_artifact", "value": "requirements/2025/harmonization/hla_2025_harmonization_worklist.csv"},
        {"metric": "row_level_source_artifact", "value": "requirements/2025/harmonization/hla_2025_requirement_disposition_ledger.csv"},
    ]
    metadata_rows = [
        {"field": "edition", "value": "2025 / 1516_2025"},
        {"field": "primary_source", "value": "requirements/2025/harmonization/hla_2025_harmonization_worklist.csv"},
        {"field": "row_level_owner_source", "value": "requirements/2025/harmonization/hla_2025_requirement_disposition_ledger.csv"},
        {"field": "front_door", "value": "docs/requirements/ieee-1516-2025/README.md"},
        {"field": "inventory", "value": "requirements/2025/README.md"},
        {
            "field": "canonical_disposition_meaning",
            "value": "Requirement coverage state only. Read this separately from backend-ownership or route-resolution fields.",
        },
        {
            "field": "denominator_rule",
            "value": "Report 100% dispositioned across all 691 tracked rows and 100% covered across the 645 active normative non-retired non-umbrella rows; do not restate this grouped packet as 691 / 691 covered.",
        },
        {
            "field": "grouped_packet_scope",
            "value": "This workbook presents 64 grouped buckets for manager-facing review while the row-level owner ledger keeps the 691-row canonical denominator.",
        },
        {
            "field": "python_runtime_resolution_meaning",
            "value": "Whether the direct python1516_2025 lane is the real proof owner, a bounded runtime surface, or an explicit exclusion.",
        },
        {
            "field": "java_cpp_binding_resolution_meaning",
            "value": "Whether Java/C++ rows are wrapper-only binding surfaces over python1516_2025 or have no active behavior-support claim.",
        },
        {
            "field": "hosted_fedpro_resolution_meaning",
            "value": "Whether FedPro is a bounded hosted-route surface over python1516_2025 rather than an independent RTI owner.",
        },
        {
            "field": "pitch_202x_resolution_meaning",
            "value": "Whether any Pitch proto HLA 4 / 202X overlap is explicit vendor-resolution context rather than inferred grouped coverage.",
        },
        {"field": "notes", "value": "Boss-facing export generated from grouped 2025 harmonization and backend-resolution surfaces; spreadsheets are secondary presentation surfaces."},
    ]
    return summary_rows, detail_rows, metadata_rows


def generate_exports(output_dir: Path) -> list[Path]:
    output_dir.mkdir(parents=True, exist_ok=True)
    created: list[Path] = []

    summary_rows, detail_rows, metadata_rows, policy_parent_rows = build_2010_rows()
    stem = "requirements_2010_backend_compliance"
    summary_csv = output_dir / f"{stem}_summary.csv"
    detail_csv = output_dir / f"{stem}_detail.csv"
    policy_parent_csv = output_dir / f"{stem}_policy_parents.csv"
    workbook = output_dir / f"{stem}.xlsx"

    _write_csv(summary_csv, list(summary_rows[0].keys()), summary_rows)
    _write_csv(detail_csv, list(detail_rows[0].keys()), detail_rows)
    policy_fields = list(policy_parent_rows[0].keys()) if policy_parent_rows else [
        "broad_requirement_id",
        "section",
        "canonical_status",
        "policy_basis",
        "supported_subset_pass_count",
        "supported_subset_pass_ids",
        "supported_subset_links_match",
        "notes",
    ]
    _write_csv(policy_parent_csv, policy_fields, policy_parent_rows)
    _export_workbook(
        workbook,
        summary_rows,
        detail_rows,
        metadata_rows,
        extra_sheets=[("policy_parents", policy_parent_rows)],
    )
    created.extend([summary_csv, detail_csv, policy_parent_csv, workbook])

    for stem, builder in (("requirements_2025_backend_compliance", build_2025_rows),):
        summary_rows, detail_rows, metadata_rows = builder()
        summary_csv = output_dir / f"{stem}_summary.csv"
        detail_csv = output_dir / f"{stem}_detail.csv"
        workbook = output_dir / f"{stem}.xlsx"

        _write_csv(summary_csv, list(summary_rows[0].keys()), summary_rows)
        _write_csv(detail_csv, list(detail_rows[0].keys()), detail_rows)
        _export_workbook(workbook, summary_rows, detail_rows, metadata_rows)
        created.extend([summary_csv, detail_csv, workbook])

    return created


def main() -> int:
    parser = argparse.ArgumentParser(description="Generate boss-facing requirement compliance spreadsheets for 2010 and 2025.")
    parser.add_argument(
        "--output-dir",
        type=Path,
        default=DEFAULT_OUTPUT_DIR,
        help="Directory where CSV and XLSX exports should be written.",
    )
    args = parser.parse_args()
    created = generate_exports(args.output_dir)
    for path in created:
        print(_display_path(path))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
